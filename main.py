import hashlib
import hmac
import io
import json
import os
import re
import secrets
import sqlite3
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import pdfplumber
from fastapi import FastAPI, File, Form, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from pydantic import BaseModel, EmailStr, Field

# (opcional) para hacer el logo semi-transparente tipo watermark
try:
    from PIL import Image as PILImage

    PIL_OK = True
except Exception:
    PIL_OK = False


app = FastAPI(title="Extractos PDF → Excel")

DEFAULT_ALLOWED_ORIGINS = [
    "http://127.0.0.1:5500",
    "http://localhost:5500",
    "https://miamilab.ai",
    "https://www.miamilab.ai",
]


def _allowed_origins() -> List[str]:
    env = os.getenv("ALLOWED_ORIGINS", "").strip()
    if env:
        return [o.strip() for o in env.split(",") if o.strip()]
    return DEFAULT_ALLOWED_ORIGINS


app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins(),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# =========================
# AUTH + PAGO
# =========================
USERS_PATH = Path("users.json")
DB_PATH = Path(os.getenv("EXTRACTO_DB_PATH", "extracto.db"))
SESSION_TTL_HOURS = int(os.getenv("SESSION_TTL_HOURS", "24"))
DEBUG_MAIN_FILE_ENABLED = os.getenv("ENABLE_MAIN_FILE_ENDPOINT", "false").lower() == "true"
ADMIN_API_KEY = os.getenv("ADMIN_API_KEY", "").strip()


class Credentials(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6, max_length=128)


class PayRequest(BaseModel):
    amount: float = Field(gt=0)


def _db_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _init_db() -> None:
    with _db_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                email TEXT PRIMARY KEY,
                salt TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                is_paid INTEGER NOT NULL DEFAULT 0,
                last_payment_amount REAL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS sessions (
                token TEXT PRIMARY KEY,
                email TEXT NOT NULL,
                created_at TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                FOREIGN KEY(email) REFERENCES users(email) ON DELETE CASCADE
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_sessions_email ON sessions(email)")


def _migrate_users_json_if_needed() -> None:
    if not USERS_PATH.exists():
        return

    with _db_connection() as conn:
        has_users = conn.execute("SELECT 1 FROM users LIMIT 1").fetchone() is not None
        if has_users:
            return

        try:
            users = json.loads(USERS_PATH.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return

        now = datetime.now(timezone.utc).isoformat()
        for email, data in users.items():
            conn.execute(
                """
                INSERT OR IGNORE INTO users(email, salt, password_hash, is_paid, last_payment_amount, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    email,
                    data.get("salt", ""),
                    data.get("password_hash", ""),
                    1 if data.get("is_paid") else 0,
                    data.get("last_payment_amount"),
                    now,
                    now,
                ),
            )


def _user_row_to_dict(row: sqlite3.Row) -> dict:
    return {
        "email": row["email"],
        "salt": row["salt"],
        "password_hash": row["password_hash"],
        "is_paid": bool(row["is_paid"]),
        "last_payment_amount": row["last_payment_amount"],
    }


def _get_user_by_email(email: str) -> Optional[dict]:
    with _db_connection() as conn:
        row = conn.execute(
            "SELECT email, salt, password_hash, is_paid, last_payment_amount FROM users WHERE email = ?",
            (email,),
        ).fetchone()
    return _user_row_to_dict(row) if row else None


def _create_user(email: str, salt: str, password_hash: str) -> None:
    now = datetime.now(timezone.utc).isoformat()
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT INTO users(email, salt, password_hash, is_paid, last_payment_amount, created_at, updated_at)
            VALUES (?, ?, ?, 0, NULL, ?, ?)
            """,
            (email, salt, password_hash, now, now),
        )


def _create_session(email: str) -> str:
    token = secrets.token_urlsafe(32)
    now = datetime.now(timezone.utc)
    expires_at = now + timedelta(hours=SESSION_TTL_HOURS)
    with _db_connection() as conn:
        conn.execute(
            "INSERT INTO sessions(token, email, created_at, expires_at) VALUES (?, ?, ?, ?)",
            (token, email, now.isoformat(), expires_at.isoformat()),
        )
    return token


def _delete_session(token: str) -> None:
    with _db_connection() as conn:
        conn.execute("DELETE FROM sessions WHERE token = ?", (token,))


def _get_email_from_session(token: str) -> Optional[str]:
    with _db_connection() as conn:
        row = conn.execute("SELECT email, expires_at FROM sessions WHERE token = ?", (token,)).fetchone()
        if not row:
            return None
        if datetime.fromisoformat(row["expires_at"]) < datetime.now(timezone.utc):
            conn.execute("DELETE FROM sessions WHERE token = ?", (token,))
            return None
        return row["email"]


def _set_user_payment(email: str, amount: float) -> None:
    now = datetime.now(timezone.utc).isoformat()
    with _db_connection() as conn:
        conn.execute(
            "UPDATE users SET is_paid = 1, last_payment_amount = ?, updated_at = ? WHERE email = ?",
            (amount, now, email),
        )


def _hash_password(password: str, salt: str) -> str:
    hashed = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120_000)
    return hashed.hex()


def _create_password_record(password: str) -> Dict[str, str]:
    salt = secrets.token_hex(16)
    return {"salt": salt, "hash": _hash_password(password, salt)}


def _verify_password(password: str, salt: str, expected_hash: str) -> bool:
    computed = _hash_password(password, salt)
    return hmac.compare_digest(computed, expected_hash)


def _extract_bearer_token(authorization: Optional[str]) -> str:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Debes iniciar sesión.")
    token = authorization.split(" ", 1)[1].strip()
    if not token:
        raise HTTPException(status_code=401, detail="Token inválido.")
    return token


def _require_admin_key(x_admin_key: Optional[str]) -> None:
    if not ADMIN_API_KEY:
        raise HTTPException(status_code=404, detail="No encontrado.")
    if not x_admin_key or not hmac.compare_digest(x_admin_key, ADMIN_API_KEY):
        raise HTTPException(status_code=403, detail="No autorizado.")


def _get_current_user(authorization: Optional[str]) -> dict:
    token = _extract_bearer_token(authorization)
    user_email = _get_email_from_session(token)
    if not user_email:
        raise HTTPException(status_code=401, detail="Sesión expirada o inválida.")

    user = _get_user_by_email(user_email)
    if not user:
        raise HTTPException(status_code=401, detail="Usuario no encontrado.")
    return user


def _require_paid_user(authorization: Optional[str]) -> dict:
    user = _get_current_user(authorization)
    if not user.get("is_paid"):
        raise HTTPException(status_code=402, detail="Necesitás pagar para convertir archivos.")
    return user


@app.post("/auth/register")
def register_user(payload: Credentials):
    email = payload.email.lower()

    if _get_user_by_email(email):
        raise HTTPException(status_code=409, detail="Ese usuario ya existe.")

    password_record = _create_password_record(payload.password)
    _create_user(email, password_record["salt"], password_record["hash"])

    return {"message": "Usuario creado correctamente."}


@app.post("/auth/login")
def login_user(payload: Credentials):
    email = payload.email.lower()
    user = _get_user_by_email(email)
    if not user:
        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos.")

    if not _verify_password(payload.password, user["salt"], user["password_hash"]):
        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos.")

    token = _create_session(email)

    return {
        "token": token,
        "email": email,
        "is_paid": bool(user.get("is_paid")),
    }


@app.get("/auth/me")
def auth_me(authorization: Optional[str] = Header(default=None)):
    user = _get_current_user(authorization)
    return {
        "email": user["email"],
        "is_paid": bool(user.get("is_paid")),
        "last_payment_amount": user.get("last_payment_amount"),
    }


@app.post("/billing/pay")
def make_payment(payload: PayRequest, authorization: Optional[str] = Header(default=None)):
    user = _get_current_user(authorization)
    email = user["email"]

    _set_user_payment(email, payload.amount)

    return {
        "message": "Pago aprobado. Ya podés convertir extractos.",
        "email": email,
        "is_paid": True,
        "amount": payload.amount,
    }


@app.post("/auth/logout")
def logout_user(authorization: Optional[str] = Header(default=None)):
    token = _extract_bearer_token(authorization)
    _delete_session(token)
    return {"message": "Sesión cerrada."}


_init_db()
_migrate_users_json_if_needed()


# =========================
# CONFIG EXCEL
# =========================
_BASE_DIR = Path(__file__).resolve().parent


def _resolve_logo_path() -> str:
    for filename in ("Isologo mIAmi Lab_cyan.png", "Logo mIAmi Lab.png"):
        candidate = _BASE_DIR / filename
        if candidate.exists():
            return str(candidate)
    return str(_BASE_DIR / "Logo mIAmi Lab.png")


LOGO_PATH = _resolve_logo_path()

# Filas de “encabezado” arriba de la tabla
META_ROWS = 3  # Cliente / Banco / Powered
BLANK_ROWS = 1  # una fila en blanco
PANDAS_STARTROW = META_ROWS + BLANK_ROWS  # 0-index (pandas)
HEADER_ROW_EXCEL = PANDAS_STARTROW + 1  # 1-index (excel): header de la tabla
DATA_START_ROW = HEADER_ROW_EXCEL + 1  # primera fila con datos

# Columnas (SIN referencia):
# A fecha, B descripcion, C debito, D credito, E saldo, F control, G control_diff, H categoria
COL_DEB = "C"
COL_CRE = "D"
COL_SAL = "E"
COL_CTRL = "F"
COL_DIFF = "G"


# =========================
# REGEX
# =========================
DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{2,4}$")
DATE_TOKEN_RE = re.compile(r"^\d{2}/\d{2}/\d{2,4}\s+")
MONEY_RE = re.compile(r"-?(?:\d{1,3}\.)*\d{1,3},\d{2}")
DATE_ANY_RE = re.compile(r"\b(\d{2}/\d{2}/\d{2,4})\b")

UNDERSCORE_RE = re.compile(r"^_+$")
MONEY_TOKEN_RE = re.compile(r"(?:-?(?:\d{1,3}\.)*\d{1,3},\d{2}-?)")


# =========================
# FOOTERS / LEYENDAS
# =========================
FOOTER_PHRASES = (
    "Los depósitos en pesos y en moneda extranjera",
    "el total de la garantía por persona y por depósito",
    "Ley 24.485",
    "Decreto",
    "Com. “A” 2337",
    'Com. "A" 2337',
    # --- NACIÓN / leyendas que se cuelan en descripción ---
    'SOLICITAR LA "CAJA',
    "SOLICITAR LA CAJA",
    'CONSULTAR EL "REGIMEN',
    "CONSULTAR EL REGIMEN",
    "REGIMEN SOBRE LA BASE",
    "SE NOS FORMULE",
    "SISTEMA DEL BANCO",
    "ART. OPERATIVAS",
    "CARACTERISTICAS",
    "CARACTERÍSTICAS",
    "FORMULE",
)


# =========================
# CATEGORÍAS (prioridad importa)
# =========================
CATEGORY_RULES = [
    ("SALDO INICIAL", "Saldo inicial"),
    ("DBCR", "IMPUESTO AL DÉBITO Y CRÉDITO"),
    ("TASA GRAL", "IMPUESTO AL DÉBITO Y CRÉDITO"),
    ("GRAVAMEN LEY 25413", "IMPUESTO AL DÉBITO Y CRÉDITO"),
    ("LEY 25413", "IMPUESTO AL DÉBITO Y CRÉDITO"),
    ("25413", "IMPUESTO AL DÉBITO Y CRÉDITO"),
    ("LEY 25.413", "IMPUESTO AL DÉBITO Y CRÉDITO"),
    ("25.413", "IMPUESTO AL DÉBITO Y CRÉDITO"),
    ("0,6%", "IMPUESTO AL DÉBITO Y CRÉDITO"),
    ("DEBITO 0,6", "IMPUESTO AL DÉBITO Y CRÉDITO"),
    ("DÉBITO 0,6", "IMPUESTO AL DÉBITO Y CRÉDITO"),
    # IVA percepciones/retenciones (antes que IVA débito)
    ("RETEN. I.V.A", "IVA PERCEPCIONES"),
    ("RETEN I.V.A", "IVA PERCEPCIONES"),
    ("RETEN. IVA", "IVA PERCEPCIONES"),
    ("RETEN IVA", "IVA PERCEPCIONES"),
    ("RETEN", "IVA PERCEPCIONES"),
    ("RETENC", "IVA PERCEPCIONES"),
    ("IVA PERCEP", "IVA PERCEPCIONES"),
    ("IVA PERCEPC", "IVA PERCEPCIONES"),
    ("PERCEP IVA", "IVA PERCEPCIONES"),
    ("RG.2408", "IVA PERCEPCIONES"),
    ("RG 2408", "IVA PERCEPCIONES"),
    ("2408", "IVA PERCEPCIONES"),
    ("IVA CRED", "IVA CRÉDITO"),
    ("CREDITO FISCAL", "IVA CRÉDITO"),
    ("IVA 21%", "IVA DÉBITO"),
    ("IVA 21", "IVA DÉBITO"),
    ("IVA 10,5", "IVA DÉBITO"),
    ("I.V.A. BASE", "IVA DÉBITO"),
    ("IVA BASE", "IVA DÉBITO"),
    ("IVA DEB", "IVA DÉBITO"),
    ("DEBITO FISCAL", "IVA DÉBITO"),
    ("I.V.A.", "IVA DÉBITO"),
    ("SIRCREB", "SIRCREB"),
    ("AFIP", "AFIP"),
    # LIQUIDACIONES antes que DEPÓSITOS
    ("ACRED.LIQ.TC", "LIQUIDACION"),
    ("ACRED. LIQ. TC", "LIQUIDACION"),
    ("ACRED LIQ", "LIQUIDACION"),
    ("LIQ.TC", "LIQUIDACION"),
    ("LIQ TC", "LIQUIDACION"),
    ("LIQUIDACION", "LIQUIDACION"),
    ("LIQ", "LIQUIDACION"),
    ("FISERV", "LIQUIDACION"),
    ("PRISMA", "LIQUIDACION"),
    ("VISA", "LIQUIDACION"),
    ("MASTERCARD", "LIQUIDACION"),
    ("AMEX", "LIQUIDACION"),
    ("ACREDITACION", "DEPÓSITOS"),
    ("ACRED.", "DEPÓSITOS"),
    ("ACRED ", "DEPÓSITOS"),
    ("DEPOSITO", "DEPÓSITOS"),
    ("DEPOS", "DEPÓSITOS"),
    ("CANJE INTERNO", "DEPÓSITOS"),
    ("COMISION", "COMISIONES Y GASTOS BANCARIOS"),
    ("COMISIONES", "COMISIONES Y GASTOS BANCARIOS"),
    ("SERVICIO DE CUENTA", "COMISIONES Y GASTOS BANCARIOS"),
    ("PAQUETES", "COMISIONES Y GASTOS BANCARIOS"),
    ("GASTO", "COMISIONES Y GASTOS BANCARIOS"),
    ("ADM.VALORES", "COMISIONES Y GASTOS BANCARIOS"),
    ("VALORES AL COBRO", "COMISIONES Y GASTOS BANCARIOS"),
    ("CHEQUE", "CHEQUE"),
    ("CHEQ.", "CHEQUE"),
    ("CHEQ", "CHEQUE"),
    ("CHQ", "CHEQUE"),
    ("RENTAS", "DGR"),
    ("DIRECCION GENERAL DE RENTAS", "DGR"),
    ("DIR GRL DE RENTAS", "DGR"),
    ("DB PM/TOT", "TRANSFERENCIA"),
    ("DB PM/TOTCORPT", "TRANSFERENCIA"),
    ("TOTCORPT", "TRANSFERENCIA"),
    ("TRANSFER", "TRANSFERENCIA"),
    ("TRANSF", "TRANSFERENCIA"),
    ("TRF", "TRANSFERENCIA"),
    ("TEF", "TRANSFERENCIA"),
    ("MACRONLINE", "TRANSFERENCIA"),
    ("TRANSFISC", "TRANSFERENCIA"),
]


def classify(desc: str) -> str:
    d = (desc or "").upper().strip()
    if d == "SALDO INICIAL":
        return "Saldo inicial"
    for k, cat in CATEGORY_RULES:
        if k in d:
            return cat
    return "OTROS"


# =========================
# HELPERS
# =========================
def to_float_ar(s: str) -> Optional[float]:
    if not s:
        return None

    raw = str(s).strip().replace("\u00a0", " ")

    neg = False
    if raw.endswith("-"):
        neg = True
        raw = raw[:-1].strip()

    if raw.startswith("(") and raw.endswith(")"):
        neg = True
        raw = raw[1:-1].strip()

    raw = re.sub(r"[^\d\.\,\-]", "", raw)

    m = MONEY_RE.search(raw)
    if not m:
        return None

    num = m.group(0)

    if num.startswith("-"):
        return float(num.replace(".", "").replace(",", "."))

    val = float(num.replace(".", "").replace(",", "."))
    return -val if neg else val


# =========================
# CLIENTE DESDE PDF (MEJORADO)
# =========================
CLIENT_BLACKLIST = [
    "BANCO",
    "SANTANDER",
    "MACRO",
    "NACION",
    "NACIÓN",
    "EXTRACTO",
    "RESUMEN",
    "CUENTA",
    "CBU",
    "ALIAS",
    "SUCURSAL",
    "DOMICILIO",
    "CUIT",
    "C.U.I.T",
    "CUIL",
    "IVA",
    "PERIODO",
    "PERÍODO",
    "HOJA",
    "EMISION",
    "EMISIÓN",
]


def _clean_client_candidate(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    s = s.strip("-:·| ")
    s = re.sub(r"\bDESDE\s*:\s*\d{2}/\d{2}/\d{2,4}\b.*$", "", s, flags=re.I).strip()
    s = re.sub(r"\bHASTA\s*:\s*\d{2}/\d{2}/\d{2,4}\b.*$", "", s, flags=re.I).strip()
    return s


def _looks_like_address(s: str) -> bool:
    s2 = (s or "").strip()
    if re.search(r"\b\d{1,5}\b", s2) and len(s2.split()) <= 4:
        return True
    return False


def _is_good_name(s: str) -> bool:
    if not s:
        return False
    up = norm_ascii(s)

    if any(b in up for b in CLIENT_BLACKLIST):
        return False

    if not re.search(r"[A-ZÁÉÍÓÚÑ]", s.upper()):
        return False

    if _looks_like_address(s):
        return False

    if len(s.split()) < 2:
        return False

    return True


def extract_client_name(pdf_bytes: bytes) -> Optional[str]:
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages[:2]:
                text = page.extract_text() or ""
                if not text:
                    continue

                raw_lines = [l.strip() for l in text.split("\n") if l.strip()]

                for line in raw_lines[:60]:
                    line_clean = _clean_client_candidate(line)
                    up = norm_ascii(line_clean)

                    if "CUIT" in up or "C.U.I.T" in up:
                        m = re.search(r"\b(C\.?U\.?I\.?T)\b", up)
                        if m:
                            before = line_clean[: m.start()].strip(" -:|")
                            before = _clean_client_candidate(before)

                            if _is_good_name(before):
                                before = re.sub(r"\bDESDE\b.*$", "", before, flags=re.I).strip()
                                return before.title() if before.isupper() else before

                for line in raw_lines[:20]:
                    line_clean = _clean_client_candidate(line)
                    if re.search(r"\bDESDE\s*:\s*\d{2}/\d{2}/\d{2,4}\b", line_clean, flags=re.I):
                        cand = re.sub(r"\bDESDE\b.*$", "", line_clean, flags=re.I).strip()
                        if _is_good_name(cand):
                            return cand.title() if cand.isupper() else cand

                for line in raw_lines[:25]:
                    cand = _clean_client_candidate(line)
                    if cand.isupper() and _is_good_name(cand):
                        return cand.title()
    except Exception:
        pass

    return None


def is_footer_or_noise(line: str) -> bool:
    if not line:
        return True
    s = line.strip()
    if s.isdigit() and len(s) <= 3:
        return True
    for p in FOOTER_PHRASES:
        if p.lower() in s.lower():
            return True
    return False


def group_lines(words, tol=8):
    words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    lines, cur, last_y = [], [], None
    for w in words:
        y = w["top"]
        if last_y is None or abs(y - last_y) <= tol:
            cur.append(w)
            last_y = y if last_y is None else (last_y + y) / 2
        else:
            lines.append(cur)
            cur = [w]
            last_y = y
    if cur:
        lines.append(cur)
    return lines


def group_lines_simple(words, tol=8):
    words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    lines, cur, last_y = [], [], None
    for w in words:
        y = w["top"]
        if last_y is None or abs(y - last_y) <= tol:
            cur.append(w)
            last_y = y
        else:
            lines.append(cur)
            cur = [w]
            last_y = y
    if cur:
        lines.append(cur)
    return lines


def line_text(lw):
    return " ".join(w["text"] for w in sorted(lw, key=lambda w: w["x0"])).strip()


def split_by_bounds(lw, bounds):
    cols = [""] * (len(bounds) - 1)
    for w in sorted(lw, key=lambda w: w["x0"]):
        for i in range(len(bounds) - 1):
            if bounds[i] <= w["x0"] < bounds[i + 1]:
                cols[i] = (cols[i] + " " + w["text"]).strip()
                break
    return cols


def clean_desc(desc: str) -> str:
    desc = (desc or "").strip()
    desc = DATE_TOKEN_RE.sub("", desc).strip()
    desc = re.sub(r"\s+", " ", desc).strip()
    return desc


def strip_legal_tail(desc: str) -> str:
    if not desc:
        return desc

    up = desc.upper()

    hard_triggers = [
        'SOLICITAR LA "CAJA',
        "SOLICITAR LA CAJA",
        'CONSULTAR EL "REGIMEN',
        "CONSULTAR EL REGIMEN",
        "REGIMEN SOBRE LA BASE",
        "SE NOS FORMULE",
        "SISTEMA DEL BANCO",
        "ART. OPERATIVAS",
        "CARACTERISTICAS",
        "CARACTERÍSTICAS",
        "FORMULE",
        " DEL SISTEMA",
        " EN SU DEL DEFECTO",
        " EN SU DEFECTO",
        " SE DEL BANCO",
        " OPERATIVAS, BANCO",
        " OPERATIVAS, EL OPERACIONES",
        " LAS NORMAS",
        " BANCO CENTRAL",
        " WWW.BCRA.GOB.AR",
        " HTTP://",
    ]

    cut_pos = None
    for t in hard_triggers:
        p = up.find(t)
        if p != -1:
            cut_pos = p if cut_pos is None else min(cut_pos, p)

    p_sobre = up.find(" SOBRE ")
    if p_sobre != -1:
        tail = up[p_sobre:]
        if any(sig in tail for sig in ["DEPÓSIT", "DEPOSIT", "REGIMEN", "RÉGIMEN", "ART.", "BANCO"]):
            cut_pos = p_sobre if cut_pos is None else min(cut_pos, p_sobre)

    if cut_pos is not None:
        desc = desc[:cut_pos].strip()

    return re.sub(r"\s+", " ", desc).strip()


def norm_ascii(s: str) -> str:
    return (
        (s or "")
        .upper()
        .replace("Ó", "O")
        .replace("É", "E")
        .replace("Í", "I")
        .replace("Á", "A")
        .replace("Ú", "U")
        .replace("Ñ", "N")
    )


def extract_date_any(*parts: str) -> str:
    text = " ".join([p or "" for p in parts])
    m = DATE_ANY_RE.search(text)
    return m.group(1) if m else ""


def _extract_money_tokens(text: str) -> List[str]:
    if not text:
        return []
    return MONEY_TOKEN_RE.findall(str(text))


def _classify_one_amount_nacion(
    desc: str, amt: float, saldo: float, prev_saldo: Optional[float]
) -> Tuple[Optional[float], Optional[float]]:
    d = (desc or "").upper()

    if "S/DEB" in d or " DEB" in d or d.endswith("DEB"):
        return amt, None
    if "S/CRED" in d or " CRED" in d or "ACRED" in d:
        return None, amt

    if prev_saldo is not None:
        if saldo > prev_saldo:
            return None, amt
        if saldo < prev_saldo:
            return amt, None

    return None, None


def _parse_amounts_from_full_row_nacion(
    desc: str,
    c0: str,
    c1: str,
    c2: str,
    c3: str,
    c4: str,
    c5: str,
    prev_saldo: Optional[float],
) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    full = " ".join([c0 or "", c1 or "", c2 or "", c3 or "", c4 or "", c5 or ""]).strip()
    tokens = _extract_money_tokens(full)
    if len(tokens) < 1:
        return None, None, None

    saldo = to_float_ar(tokens[-1])
    if saldo is None:
        return None, None, None

    mov_tokens = tokens[:-1]
    mov_amounts = [to_float_ar(t) for t in mov_tokens]
    mov_amounts = [x for x in mov_amounts if x is not None]

    if len(mov_amounts) == 0:
        return None, None, float(saldo)

    if len(mov_amounts) == 1:
        deb, cre = _classify_one_amount_nacion(desc, float(mov_amounts[0]), float(saldo), prev_saldo)
        return deb, cre, float(saldo)

    a = float(mov_amounts[-2])
    b = float(mov_amounts[-1])

    if prev_saldo is None:
        return a, b, float(saldo)

    delta = float(saldo) - float(prev_saldo)
    net1 = b - a
    net2 = a - b
    if abs(net1 - delta) <= abs(net2 - delta):
        return a, b, float(saldo)
    return b, a, float(saldo)


# =========================
# HELPERS PARA UNIFICAR
# =========================
def parse_fecha_ddmmaa(s: str):
    if not s:
        return pd.NaT
    s = str(s).strip()
    fmt = "%d/%m/%y" if len(s.split("/")[-1]) == 2 else "%d/%m/%Y"
    return pd.to_datetime(s, format=fmt, errors="coerce")


def split_saldo_inicial(df: pd.DataFrame):
    if df is None or df.empty:
        return None, df

    mask = df["descripcion"].fillna("").str.upper().str.strip().eq("SALDO INICIAL")
    if mask.any():
        first_row = df.loc[mask].iloc[0]
        saldo_ini = first_row.get("saldo", None)
        df2 = df.loc[~mask].copy()
        return saldo_ini, df2

    saldo_ini = None
    if "saldo" in df.columns and df["saldo"].notna().any():
        saldo_ini = float(df["saldo"].dropna().iloc[0])
    return saldo_ini, df.copy()


def merge_extracts_keep_first_saldo(dfs: List[pd.DataFrame]) -> pd.DataFrame:
    valid = [df for df in (dfs or []) if df is not None and not df.empty]
    if not valid:
        return pd.DataFrame(
            columns=["fecha", "descripcion", "debito", "credito", "saldo", "control", "control_diff", "categoria"]
        )

    df_infos = []
    for idx, df in enumerate(valid):
        tmp = df.copy()
        tmp["_fecha_dt"] = tmp["fecha"].apply(parse_fecha_ddmmaa)
        min_dt = tmp["_fecha_dt"].min()
        df_infos.append((min_dt, idx, df))
    df_infos.sort(key=lambda x: (pd.isna(x[0]), x[0], x[1]))
    ordered_dfs = [x[2] for x in df_infos]

    parts = []
    saldo_candidates = []

    for file_rank, df in enumerate(ordered_dfs):
        saldo_ini, df_wo = split_saldo_inicial(df)

        df_wo = df_wo.copy().reset_index(drop=True)
        df_wo["_file_rank"] = file_rank
        df_wo["_row_idx"] = range(len(df_wo))

        fechas_dt = df_wo["fecha"].apply(parse_fecha_ddmmaa) if not df_wo.empty else pd.Series([pd.NaT])
        min_dt = fechas_dt.min()

        if saldo_ini is not None:
            saldo_candidates.append((min_dt, float(saldo_ini)))

        parts.append(df_wo)

    merged = pd.concat(parts, ignore_index=True)
    merged["_fecha_dt"] = merged["fecha"].apply(parse_fecha_ddmmaa)

    merged = (
        merged.sort_values(["_fecha_dt", "_file_rank", "_row_idx"], kind="mergesort")
        .drop(columns=["_fecha_dt", "_file_rank", "_row_idx"])
        .reset_index(drop=True)
    )

    saldo_candidates = [x for x in saldo_candidates if pd.notna(x[0])]
    if saldo_candidates:
        saldo_candidates.sort(key=lambda x: x[0])
        saldo_ini_final = saldo_candidates[0][1]
    else:
        saldo_ini_final = float(merged.iloc[0]["saldo"]) if not merged.empty else 0.0

    fecha_ini_final = merged.iloc[0]["fecha"] if not merged.empty else ""

    fila_ini = pd.DataFrame(
        [
            {
                "fecha": fecha_ini_final,
                "descripcion": "SALDO INICIAL",
                "debito": None,
                "credito": float(saldo_ini_final),
                "saldo": float(saldo_ini_final),
                "control": None,
                "control_diff": None,
                "categoria": "Saldo inicial",
            }
        ]
    )

    merged = pd.concat([fila_ini, merged], ignore_index=True)
    merged["categoria"] = merged["descripcion"].apply(classify)
    merged["control"] = None
    merged["control_diff"] = None

    return merged[["fecha", "descripcion", "debito", "credito", "saldo", "control", "control_diff", "categoria"]]


# =========================
# HEADER BOUNDS (MACRO)
# =========================
def find_header_bounds_macro(lines) -> Tuple[Optional[List[float]], Optional[float]]:
    for lw in lines:
        t = norm_ascii(line_text(lw))
        if all(k in t for k in ["FECHA", "DESCRIPCION", "REFERENCIA", "DEBITOS", "CREDITOS", "SALDO"]):
            xs = {norm_ascii(w["text"]): w["x0"] for w in lw}

            def get_x(token: str) -> Optional[float]:
                token = norm_ascii(token)
                if token in xs:
                    return xs[token]
                for w in lw:
                    wt = norm_ascii(w["text"])
                    if wt.startswith(token[:5]):
                        return w["x0"]
                return None

            x_fecha = get_x("FECHA")
            x_desc = get_x("DESCRIPCION")
            x_ref = get_x("REFERENCIA")
            x_deb = get_x("DEBITOS")
            x_cre = get_x("CREDITOS")
            x_sal = get_x("SALDO")

            if any(v is None for v in [x_fecha, x_desc, x_ref, x_deb, x_cre, x_sal]):
                continue

            x = [x_fecha, x_desc, x_ref, x_deb, x_cre, x_sal]
            bounds = [0.0]
            bounds.append((x[0] + x[1]) / 2)
            bounds.append(((x[1] + x[2]) / 2) + 25)
            bounds.append((x[2] + x[3]) / 2)
            bounds.append((x[3] + x[4]) / 2)
            bounds.append((x[4] + x[5]) / 2)
            bounds.append(10_000.0)
            return bounds, lw[0]["top"]

    return None, None


# =========================
# HEADER BOUNDS (NACIÓN)
# =========================
def find_header_bounds_nacion(lines) -> Tuple[Optional[List[float]], Optional[float]]:
    header_variants = [
        ("FECHA", "DESCRIPCION", "REFERENCIA", "DEBITOS", "CREDITOS", "SALDO"),
        ("FECHA", "DESCRIPCIÓN", "REFERENCIA", "DÉBITOS", "CRÉDITOS", "SALDO"),
        ("FECHA", "MOVIMIENTOS", "COMPROB", "DEBITOS", "CREDITOS", "SALDO"),
    ]

    for lw in lines:
        t = norm_ascii(line_text(lw))
        for hv in header_variants:
            hvn = tuple(norm_ascii(x) for x in hv)
            if all(k in t for k in hvn):
                xs = {norm_ascii(w["text"]): w["x0"] for w in lw}

                def get_x(token: str) -> Optional[float]:
                    token = norm_ascii(token)
                    if token in xs:
                        return xs[token]
                    for w in lw:
                        wt = norm_ascii(w["text"])
                        if wt.startswith(token[:5]):
                            return w["x0"]
                    return None

                x_fecha = get_x("FECHA")
                x_desc = get_x("DESCRIPCION") or get_x("MOVIMIENTOS")
                x_ref = get_x("REFERENCIA") or get_x("COMPROB")
                x_deb = get_x("DEBITOS")
                x_cre = get_x("CREDITOS")
                x_sal = get_x("SALDO")

                if any(v is None for v in [x_fecha, x_desc, x_ref, x_deb, x_cre, x_sal]):
                    continue

                x = [x_fecha, x_desc, x_ref, x_deb, x_cre, x_sal]
                bounds = [0.0]
                bounds.append((x[0] + x[1]) / 2)
                bounds.append((x[1] + x[2]) / 2)
                bounds.append((x[2] + x[3]) / 2)
                bounds.append((x[3] + x[4]) / 2)
                bounds.append((x[4] + x[5]) / 2)
                bounds.append(10_000.0)
                return bounds, lw[0]["top"]

    return None, None


# =========================
# EXTRACTOR (MACRO) - SIN REFERENCIA EN SALIDA
# =========================
def extract_macro_table(pdf_bytes: bytes) -> pd.DataFrame:
    rows: List[Dict] = []
    first_date: Optional[str] = None
    saldo_inicial: Optional[float] = None

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            words_all = page.extract_words(use_text_flow=True, keep_blank_chars=False) or []
            words = [w for w in words_all if 60 < w["top"] < (page.height - 60)]
            lines = group_lines(words, tol=8)

            bounds, header_y = find_header_bounds_macro(lines)
            if bounds is None:
                continue

            for lw in lines:
                if lw[0]["top"] <= header_y + 1:
                    continue

                tline = line_text(lw)
                if not tline or is_footer_or_noise(tline):
                    continue

                cols = split_by_bounds(lw, bounds)
                c0, c1, c2, c3, c4, c5 = (
                    cols[0].strip(),
                    cols[1].strip(),
                    cols[2].strip(),
                    cols[3].strip(),
                    cols[4].strip(),
                    cols[5].strip(),
                )

                combined = norm_ascii(f"{c0} {c1} {c2}")
                if "SALDO" in combined and "ULTIMO" in combined and "EXTRACTO" in combined:
                    sal = to_float_ar(c5)
                    if sal is not None and saldo_inicial is None and abs(sal) > 0.000001:
                        saldo_inicial = float(sal)
                    continue

                fecha = c0.split()[0].strip() if c0 else ""
                if not DATE_RE.match(fecha):
                    continue

                if first_date is None:
                    first_date = fecha

                extra0 = " ".join(c0.split()[1:]).strip()
                desc = clean_desc((extra0 + " " + c1).strip())

                deb = to_float_ar(c3)
                cre = to_float_ar(c4)
                sal = to_float_ar(c5)

                rows.append({"fecha": fecha, "descripcion": desc, "debito": deb, "credito": cre, "saldo": sal})

    df = pd.DataFrame(rows)

    if saldo_inicial is None:
        raise ValueError("No se encontró 'SALDO ULTIMO EXTRACTO' con saldo inicial (columna SALDO).")

    fecha_ini = first_date or (df.iloc[0]["fecha"] if not df.empty else "")
    fila_ini = pd.DataFrame(
        [{"fecha": fecha_ini, "descripcion": "SALDO INICIAL", "debito": None, "credito": float(saldo_inicial), "saldo": float(saldo_inicial)}]
    )

    df = pd.concat([fila_ini, df], ignore_index=True)
    df["categoria"] = df["descripcion"].apply(classify)
    df["control"] = None
    df["control_diff"] = None

    return df[["fecha", "descripcion", "debito", "credito", "saldo", "control", "control_diff", "categoria"]]


# =========================
# EXTRACTOR (NACIÓN) - SIN REFERENCIA EN SALIDA
# =========================
def extract_nacion_table(pdf_bytes: bytes) -> pd.DataFrame:
    rows: List[Dict] = []
    first_date: Optional[str] = None
    saldo_inicial: Optional[float] = None
    found_any_table = False

    SALDO_KEYS = ("SALDO ANTERIOR", "SALDO INICIAL", "SALDO ANTERIOR AL", "SALDO INICIAL AL")
    TABLE_END_KEYS = ("SALDO FINAL", "TOTAL", "TOTALES")

    prev_saldo_mov: Optional[float] = None

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            words_all = page.extract_words(use_text_flow=True, keep_blank_chars=False) or []
            words_all = [w for w in words_all if not UNDERSCORE_RE.match((w.get("text") or "").strip())]
            words = [w for w in words_all if 60 < w["top"] < (page.height - 60)]
            lines = group_lines_simple(words, tol=8)

            bounds, header_y = find_header_bounds_nacion(lines)
            if bounds is None:
                continue

            found_any_table = True

            stop_y = None
            for lw in lines:
                y = lw[0]["top"]
                if y <= header_y + 1:
                    continue
                t = norm_ascii(line_text(lw))
                if any(k in t for k in TABLE_END_KEYS):
                    stop_y = y
                    break

            for lw in lines:
                y = lw[0]["top"]
                if y <= header_y + 1:
                    continue
                if stop_y is not None and y >= stop_y:
                    break

                tline = line_text(lw)
                if not tline or is_footer_or_noise(tline):
                    continue

                cols = split_by_bounds(lw, bounds)
                c0 = (cols[0] or "").strip()
                c1 = (cols[1] or "").strip()
                c2 = (cols[2] or "").strip()
                c3 = (cols[3] or "").strip()
                c4 = (cols[4] or "").strip()
                c5 = (cols[5] or "").strip()

                row_text = norm_ascii(f"{c0} {c1} {c2} {c3} {c4} {c5}")
                if any(k in row_text for k in TABLE_END_KEYS):
                    break

                if saldo_inicial is None and any(norm_ascii(k) in row_text for k in SALDO_KEYS):
                    tokens = _extract_money_tokens(f"{c0} {c1} {c2} {c3} {c4} {c5}")
                    if tokens:
                        s = to_float_ar(tokens[-1])
                        if s is not None:
                            saldo_inicial = float(s)
                            prev_saldo_mov = float(s)
                            continue

                fecha = extract_date_any(c0, c1, c2, c3, c4, c5)
                if not fecha or not DATE_RE.match(fecha):
                    if rows:
                        maybe_legal = any(k in row_text for k in ["SOLICITAR", "CONSULTAR", "REGIMEN", "BCRA", "BANCO", "ART.", "NORMAS", "SOBRE", "HTTP"])
                        has_money = bool(_extract_money_tokens(f"{c0} {c1} {c2} {c3} {c4} {c5}"))
                        if (c1 or c0) and (not has_money) and (not maybe_legal):
                            rows[-1]["descripcion"] = strip_legal_tail((rows[-1]["descripcion"] + " " + clean_desc(c1 or c0)).strip())
                    continue

                if first_date is None:
                    first_date = fecha

                desc = strip_legal_tail(clean_desc(c1))

                deb = to_float_ar(c3)
                cre = to_float_ar(c4)
                sal = to_float_ar(c5)

                if sal is None:
                    deb2, cre2, sal2 = _parse_amounts_from_full_row_nacion(desc=desc, c0=c0, c1=c1, c2=c2, c3=c3, c4=c4, c5=c5, prev_saldo=prev_saldo_mov)
                    deb, cre, sal = deb2, cre2, sal2

                if sal is None:
                    continue

                if deb is None and cre is None and prev_saldo_mov is not None:
                    delta = float(sal) - float(prev_saldo_mov)
                    if delta > 0:
                        cre = abs(delta)
                    elif delta < 0:
                        deb = abs(delta)

                rows.append({"fecha": fecha, "descripcion": desc, "debito": deb, "credito": cre, "saldo": float(sal)})
                prev_saldo_mov = float(sal)

    df = pd.DataFrame(rows)

    if not found_any_table or df.empty:
        raise ValueError("No se encontró la tabla de movimientos (no se detectó el header o no se pudo leer el contenido).")

    if saldo_inicial is None:
        if df["saldo"].notna().any():
            saldo_inicial = float(df["saldo"].dropna().iloc[0])
        else:
            raise ValueError("No se encontró saldo inicial (ni SALDO ANTERIOR/INICIAL ni saldos en tabla).")

    fecha_ini = first_date or (df.iloc[0]["fecha"] if not df.empty else "")
    fila_ini = pd.DataFrame([{"fecha": fecha_ini, "descripcion": "SALDO INICIAL", "debito": None, "credito": float(saldo_inicial), "saldo": float(saldo_inicial)}])

    df = pd.concat([fila_ini, df], ignore_index=True)
    df["categoria"] = df["descripcion"].apply(classify)
    df["control"] = None
    df["control_diff"] = None

    return df[["fecha", "descripcion", "debito", "credito", "saldo", "control", "control_diff", "categoria"]]


# =========================
# SANTANDER (TABLA POR COORDENADAS) - SIN REFERENCIA EN SALIDA
# =========================
def find_header_bounds_santander(lines) -> Tuple[Optional[List[float]], Optional[float]]:
    for lw in lines:
        t = norm_ascii(line_text(lw))

        has_fecha = "FECHA" in t
        has_mov = "MOVIMIENTO" in t
        has_deb = "DEBITO" in t
        has_cre = "CREDITO" in t
        has_saldo = ("SALDO" in t and "CUENTA" in t) or ("SALDO EN CUENTA" in t)
        has_comp = ("COMPROBANTE" in t) or ("COMPROB" in t)

        if not (has_fecha and has_mov and has_deb and has_cre and has_saldo and has_comp):
            continue

        xs = {norm_ascii(w["text"]): w["x0"] for w in lw}

        def get_x_any(starts: List[str]) -> Optional[float]:
            for s in starts:
                s = norm_ascii(s)
                if s in xs:
                    return xs[s]
            for w in lw:
                wt = norm_ascii(w["text"])
                for s in starts:
                    s2 = norm_ascii(s)
                    if wt.startswith(s2[:5]):
                        return w["x0"]
            return None

        x_fecha = get_x_any(["FECHA"])
        x_comp = get_x_any(["COMPROBANTE", "COMPROB", "COMPROB."])
        x_mov = get_x_any(["MOVIMIENTO"])
        x_deb = get_x_any(["DEBITO", "DÉBITO"])
        x_cre = get_x_any(["CREDITO", "CRÉDITO"])
        x_sal = get_x_any(["SALDO"])

        if any(v is None for v in [x_fecha, x_comp, x_mov, x_deb, x_cre, x_sal]):
            continue

        x = [x_fecha, x_comp, x_mov, x_deb, x_cre, x_sal]
        bounds = [0.0]
        for a, b in zip(x, x[1:]):
            bounds.append((a + b) / 2)
        bounds.append(10_000.0)

        return bounds, lw[0]["top"]

    return None, None


def extract_santander_table(pdf_bytes: bytes) -> pd.DataFrame:
    rows: List[Dict] = []
    saldo_inicial: Optional[float] = None
    first_date: Optional[str] = None
    current_date: Optional[str] = None
    found_any_table = False

    TABLE_END_KEYS = ("DETALLE IMPOSITIVO", "SALDO TOTAL", "SALDO FINAL")

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            words_all = page.extract_words(use_text_flow=True, keep_blank_chars=False) or []
            words = [w for w in words_all if 60 < w["top"] < (page.height - 60)]
            lines = group_lines(words, tol=8)

            bounds, header_y = find_header_bounds_santander(lines)
            if bounds is None:
                continue

            found_any_table = True

            stop_y = None
            for lw in lines:
                y = lw[0]["top"]
                if y <= header_y + 1:
                    continue
                t = norm_ascii(line_text(lw))
                if any(k in t for k in TABLE_END_KEYS):
                    stop_y = y
                    break

            for lw in lines:
                y = lw[0]["top"]
                if y <= header_y + 1:
                    continue
                if stop_y is not None and y >= stop_y:
                    break

                tline = line_text(lw)
                if not tline or is_footer_or_noise(tline):
                    continue

                cols = split_by_bounds(lw, bounds)
                c0 = (cols[0] or "").strip()
                c1 = (cols[1] or "").strip()
                c2 = (cols[2] or "").strip()
                c3 = (cols[3] or "").strip()
                c4 = (cols[4] or "").strip()
                c5 = (cols[5] or "").strip()

                row_text_all = norm_ascii(f"{c0} {c1} {c2} {c3} {c4} {c5}")
                if any(k in row_text_all for k in TABLE_END_KEYS):
                    break

                fecha = c0[:8] if len(c0) >= 8 else ""
                if fecha and DATE_RE.match(fecha):
                    current_date = fecha
                    if first_date is None:
                        first_date = fecha
                else:
                    fecha = current_date or ""

                deb = to_float_ar(c3)
                cre = to_float_ar(c4)
                sal = to_float_ar(c5)

                has_amounts = (deb is not None) or (cre is not None) or (sal is not None)
                movimiento = clean_desc(c2)

                if movimiento and "SALDO INICIAL" in norm_ascii(movimiento) and sal is not None:
                    saldo_inicial = float(sal)
                    continue

                if (not has_amounts or (deb is None and cre is None and sal is None)) and movimiento:
                    if rows:
                        rows[-1]["descripcion"] = clean_desc((rows[-1]["descripcion"] + " " + movimiento).strip())
                    continue

                if not fecha:
                    continue

                desc = movimiento or clean_desc(c1)

                rows.append(
                    {"fecha": fecha, "descripcion": desc, "debito": deb, "credito": cre, "saldo": float(sal) if sal is not None else None}
                )

    if not found_any_table or not rows:
        raise ValueError("Santander: no se encontró la tabla de movimientos o está vacía.")

    if saldo_inicial is None:
        first_saldo = rows[0].get("saldo", None)
        if first_saldo is None:
            raise ValueError("Santander: no se encontró 'Saldo Inicial' y no hay saldo en la primera fila.")
        saldo_inicial = float(first_saldo)

    df = pd.DataFrame(rows)

    fila_ini = pd.DataFrame([{"fecha": first_date or df.iloc[0]["fecha"], "descripcion": "SALDO INICIAL", "debito": None, "credito": float(saldo_inicial), "saldo": float(saldo_inicial)}])

    df = pd.concat([fila_ini, df], ignore_index=True)
    df["categoria"] = df["descripcion"].apply(classify)
    df["control"] = None
    df["control_diff"] = None

    return df[["fecha", "descripcion", "debito", "credito", "saldo", "control", "control_diff", "categoria"]]


def _extract_bank_dataframe(bank: str, pdf_bytes: bytes) -> Tuple[pd.DataFrame, str]:
    bank_norm = (bank or "").strip().lower()
    if bank_norm == "macro":
        return extract_macro_table(pdf_bytes), "Macro"
    if bank_norm in ("nacion", "nación", "bna", "banco nacion", "banco nación"):
        return extract_nacion_table(pdf_bytes), "Nación"
    if bank_norm == "santander":
        return extract_santander_table(pdf_bytes), "Santander"
    raise HTTPException(status_code=400, detail="Banco no soportado todavía.")


# =========================
# EXCEL HELPERS (ENCABEZADO + WATERMARK + CONTROL)
# =========================
def add_excel_header(ws, client: str, bank: str):
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    ws.merge_cells("A3:H3")

    ws["A1"] = f"CLIENTE: {client or '-'}"
    ws["A2"] = f"BANCO: {bank or '-'}"
    ws["A3"] = "POWERED by MIAMI LAB ©"

    for cell in ["A1", "A2", "A3"]:
        ws[cell].font = Font(bold=True, size=12)
        ws[cell].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[4].height = 8


def add_watermark(ws, logo_path: str = LOGO_PATH):
    try:
        if PIL_OK:
            img = PILImage.open(logo_path).convert("RGBA")
            alpha = img.split()[-1]
            alpha = alpha.point(lambda p: int(p * 0.12))
            img.putalpha(alpha)

            tmp = io.BytesIO()
            img.save(tmp, format="PNG")
            tmp.seek(0)

            xlimg = XLImage(tmp)
        else:
            xlimg = XLImage(logo_path)

        xlimg.width = 380
        xlimg.height = 380
        ws.add_image(xlimg, "C8")
    except Exception:
        pass


def apply_control_formulas(
    wb_bytes: io.BytesIO, sheet_name: str = "movimientos", data_start_row: int = DATA_START_ROW
) -> io.BytesIO:
    wb_bytes.seek(0)
    wb = load_workbook(wb_bytes)
    ws = wb[sheet_name]

    max_row = ws.max_row
    if max_row < data_start_row:
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    r0 = data_start_row
    ws[f"{COL_CTRL}{r0}"].value = f"={COL_SAL}{r0}"
    ws[f"{COL_DIFF}{r0}"].value = f"={COL_SAL}{r0}-{COL_CTRL}{r0}"

    for r in range(r0 + 1, max_row + 1):
        prev = r - 1
        ws[f"{COL_CTRL}{r}"].value = f"={COL_SAL}{prev}-{COL_DEB}{r}+{COL_CRE}{r}"
        ws[f"{COL_DIFF}{r}"].value = f"={COL_SAL}{r}-{COL_CTRL}{r}"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def build_excel(df: pd.DataFrame, client: str, bank_label: str, watermark: bool = True) -> io.BytesIO:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="movimientos", startrow=PANDAS_STARTROW)
        ws = writer.sheets["movimientos"]

        add_excel_header(ws, client=client, bank=bank_label)

        for cell in ws[HEADER_ROW_EXCEL]:
            cell.font = Font(bold=True)

        if watermark:
            add_watermark(ws, LOGO_PATH)

        ws.column_dimensions["A"].width = 12  # fecha
        ws.column_dimensions["B"].width = 45  # descripcion
        ws.column_dimensions["C"].width = 12  # debito
        ws.column_dimensions["D"].width = 12  # credito
        ws.column_dimensions["E"].width = 14  # saldo
        ws.column_dimensions["F"].width = 14  # control
        ws.column_dimensions["G"].width = 14  # control_diff
        ws.column_dimensions["H"].width = 26  # categoria

    return apply_control_formulas(out, sheet_name="movimientos", data_start_row=DATA_START_ROW)


# =========================
# API
# =========================
@app.post("/convert")
async def convert(
    bank: str = Form(...),
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(default=None),
):
    _require_paid_user(authorization)

    if file.content_type not in ("application/pdf", "application/octet-stream"):
        raise HTTPException(status_code=400, detail="Subí un PDF válido.")

    pdf_bytes = await file.read()
    client_from_pdf = extract_client_name(pdf_bytes) or "-"

    try:
        df, bank_label = _extract_bank_dataframe(bank, pdf_bytes)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo extraer el extracto. Detalle: {e}")

    xlsx = build_excel(df, client=client_from_pdf, bank_label=bank_label, watermark=True)

    filename = (file.filename or "extracto.pdf").replace(".pdf", "") + ".xlsx"
    return StreamingResponse(
        xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/convert-merge")
async def convert_merge(
    bank: str = Form(...),
    files: List[UploadFile] = File(...),
    authorization: Optional[str] = Header(default=None),
):
    _require_paid_user(authorization)

    if not files:
        raise HTTPException(status_code=400, detail="No se subieron archivos.")

    bank_norm = (bank or "").strip().lower()
    all_dfs: List[pd.DataFrame] = []

    bank_label = None
    client_name = None

    for f in files:
        if f.content_type not in ("application/pdf", "application/octet-stream"):
            raise HTTPException(status_code=400, detail=f"Archivo inválido (no PDF): {f.filename}")

        pdf_bytes = await f.read()

        if client_name is None:
            client_name = extract_client_name(pdf_bytes)

        try:
            df, bank_label = _extract_bank_dataframe(bank, pdf_bytes)
            all_dfs.append(df)
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"Error procesando {f.filename}: {e}")

    if not all_dfs:
        raise HTTPException(status_code=400, detail="No se pudo procesar ningún PDF.")

    df_all = merge_extracts_keep_first_saldo(all_dfs)

    xlsx = build_excel(df_all, client=(client_name or "-"), bank_label=(bank_label or bank), watermark=True)

    filename = f"extractos_{bank_norm}_unificados.xlsx"
    return StreamingResponse(
        xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/main-file")
def get_main_file(x_admin_key: Optional[str] = Header(default=None)):
    if not DEBUG_MAIN_FILE_ENABLED:
        raise HTTPException(status_code=404, detail="No encontrado.")
    _require_admin_key(x_admin_key)

    main_path = Path(__file__)
    return StreamingResponse(
        iter([main_path.read_text(encoding="utf-8")]),
        media_type="text/plain; charset=utf-8",
    )


@app.get("/health")
def health():
    return {"ok": True}

